from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem
from PyQt5.uic import loadUi
import sys
import openpyxl


class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel data to QtableWidget")

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        self.load_data()

    def load_data(self):
        path = "excel_viewer\list-countries-world.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)

        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        row = 0
        for value in list_values[1:]:
            col = 0
            for v in value:
                self.table_widget.setItem(row, col, QTableWidgetItem(str(v)))
                col += 1
            row += 1


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    app.exec_()
